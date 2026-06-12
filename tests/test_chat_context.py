from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from ai_reporting.chat import ReportingChatService
from ai_reporting.knowledge_db import KnowledgeDatabase
from ai_reporting.knowledge import ReportingKnowledgeBase, RuleChunk
from ai_reporting.llm import ReportingLlmClient
from ai_reporting.response_format import format_chat_response
from ai_reporting.session_store import SessionStore


class CapturingLlm:
    def __init__(self):
        self.calls = []

    def complete(self, system_prompt, context_pack, user_question):
        self.calls.append(
            {
                "system_prompt": system_prompt,
                "context_pack": context_pack,
                "user_question": user_question,
            }
        )
        return type(
            "Response",
            (),
            {
                "text": "Captured model response.",
                "provider": "test",
                "model": "test-model",
                "used_live_model": True,
            },
        )()


class StubKnowledge:
    def __init__(self, rule_chunks=None, user_chunks=None):
        self.rule_chunks = rule_chunks or []
        self.user_chunks = user_chunks or []
        self.retrieve_calls = []
        self.retrieve_user_calls = []

    def retrieve(self, query, limit=5):
        self.retrieve_calls.append((query, limit))
        return self.rule_chunks[:limit]

    def retrieve_user(self, session_id, query, limit=5):
        self.retrieve_user_calls.append((session_id, query, limit))
        return self.user_chunks[:limit]

    def ingest_user_document(self, session_id, file_id, filename, path, data, summary):
        return {"documents": 1, "chunks": 1}


class ChatContextTest(unittest.TestCase):
    def _service(self, root: Path) -> ReportingChatService:
        reference = root / "reference"
        (reference / "FASB Accounting Standard Codifications (ASC)").mkdir(parents=True)
        (reference / "SEC Regulation").mkdir(parents=True)
        db = KnowledgeDatabase(root / "knowledge.db")
        return ReportingChatService(
            SessionStore(root / "sessions"),
            ReportingKnowledgeBase(reference, root / "knowledge", db),
            ReportingLlmClient(),
        )

    def test_chat_message_persists_session_upload_and_uses_preview_context(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)

            with patch.dict("os.environ", {"OPENAI_API_KEY": ""}, clear=False):
                service = self._service(root)
                payload = service.handle_message(
                    None,
                    "What rule support applies to this contract booking?",
                    [("contract.txt", b"Customer contract with payment terms and deliverables.")],
                )

            self.assertEqual(payload["status"], "answered")
            self.assertFalse(payload["used_live_model"])
            self.assertEqual(len(payload["uploaded_files"]), 1)
            self.assertEqual(payload["uploaded_files"][0]["ingestion"]["status"], "indexed")

            session_dir = root / "sessions" / payload["session_id"]
            self.assertTrue((session_dir / "manifest.json").exists())
            self.assertTrue((session_dir / "messages.json").exists())
            self.assertEqual(len(list((session_dir / "uploads").glob("*contract.txt"))), 1)
            self.assertGreater(len(service.knowledge.retrieve_user(payload["session_id"], "payment terms", limit=3)), 0)

    def test_uploaded_files_include_ingestion_status(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                StubKnowledge(),
                CapturingLlm(),
            )
            payload = service.store_uploads(None, [("policy.txt", b"Revenue recognition policy.")])

            upload = payload["uploaded_files"][0]
            self.assertEqual(upload["ingestion"]["status"], "indexed")
            self.assertEqual(upload["ingestion"]["chunk_count"], 1)
            manifest = service.store.manifest(payload["session_id"])
            self.assertEqual(manifest["uploads"][0]["ingestion"]["status"], "indexed")

    def test_disclosure_completeness_engine_reviews_uploaded_document(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            llm = CapturingLlm()
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                StubKnowledge(),
                llm,
            )
            draft = (
                "Revenue recognition policy: The company recognizes revenue when control transfers. "
                "Revenue is disaggregated by product line. Deferred revenue is a contract liability."
            ).encode("utf-8")

            payload = service.handle_message(
                None,
                "Are all the correct disclosures included?",
                [("revenue_footnote.txt", draft)],
            )

            self.assertEqual(payload["provider"], "deterministic-tool")
            self.assertEqual(payload["model"], "disclosure-completeness")
            self.assertEqual(payload["selected_skill"]["id"], "disclosure_checklist")
            self.assertEqual(payload["disclosure_review"]["topic_id"], "revenue")
            self.assertIn("checklist_items", payload["display"])
            self.assertEqual(len(payload["artifacts"]), 1)
            self.assertEqual(llm.calls, [])

            statuses = {
                item["item_id"]: item["status"]
                for item in payload["disclosure_review"]["findings"]
            }
            self.assertEqual(statuses["revenue_policy"], "included")
            self.assertEqual(statuses["contract_balances"], "included")
            self.assertIn(statuses["performance_obligations"], {"missing", "incomplete"})

    def test_disclosure_completeness_uses_prior_session_uploads(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                StubKnowledge(),
                CapturingLlm(),
            )
            upload = service.store_uploads(
                None,
                [("cash_flow_note.txt", b"Operating activities, investing activities, and financing activities are presented.")],
            )

            payload = service.handle_message(
                upload["session_id"],
                "Are any required cash flow disclosures missing?",
            )

            self.assertEqual(payload["provider"], "deterministic-tool")
            self.assertEqual(payload["disclosure_review"]["topic_id"], "cash_flow")
            self.assertEqual(payload["uploaded_files"], [])
            self.assertGreaterEqual(payload["disclosure_review"]["counts"]["missing"], 1)

    def test_amortization_request_asks_for_principal_and_rate(self) -> None:
        with tempfile.TemporaryDirectory() as directory, patch.dict("os.environ", {"OPENAI_API_KEY": ""}, clear=False):
            service = self._service(Path(directory))
            payload = service.handle_message(
                None,
                "please generate an amortization schedule with a fixed rate over 3 years",
            )

        self.assertIn("principal amount", payload["answer"])
        self.assertIn("fixed annual interest rate", payload["answer"])
        self.assertEqual(payload["provider"], "deterministic-tool")
        self.assertNotIn("payment frequency", payload["answer"].lower())
        self.assertFalse(payload.get("artifacts"))

    def test_amortization_request_generates_excel_after_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as directory, patch.dict("os.environ", {"OPENAI_API_KEY": ""}, clear=False):
            root = Path(directory)
            service = self._service(root)
            first = service.handle_message(
                None,
                "please generate an amortization schedule with a fixed rate over 3 years",
            )
            payload = service.handle_message(
                first["session_id"],
                "principal amount is $100,000 and interest rate is 6%",
            )

            self.assertEqual(payload["provider"], "deterministic-tool")
            self.assertEqual(len(payload["artifacts"]), 1)
            workbook_path = service.store.session_dir(first["session_id"]) / "artifacts" / "amortization_schedule.xlsx"
            self.assertTrue(workbook_path.exists())
            workbook = openpyxl.load_workbook(workbook_path, data_only=False)
            self.assertIn("Amortization Schedule", workbook.sheetnames)
            schedule = workbook["Amortization Schedule"]
            self.assertEqual(schedule["A1"].value, "Period")
            self.assertEqual(schedule.max_row, 39)
            self.assertEqual(schedule["A37"].value, 36)

    def test_categorized_query_adds_skill_context_to_llm_call(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            llm = CapturingLlm()
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                ReportingKnowledgeBase(root / "reference", root / "knowledge", KnowledgeDatabase(root / "knowledge.db")),
                llm,
            )
            payload = service.handle_message(None, "What does ASC 230 require for cash flow presentation?")

        self.assertEqual(payload["selected_skill"]["id"], "rule_research")
        self.assertEqual(payload["selected_skill_spec"]["id"], "rule_research")
        self.assertEqual(payload["raw_answer"], "Captured model response.")
        context = llm.calls[0]["context_pack"]
        self.assertEqual(context["selected_reporting_skill"]["id"], "rule_research")
        self.assertEqual(context["selected_skill_spec"]["id"], "rule_research")
        self.assertIn("ASC/SEC Rule Research Skill", context["selected_skill_spec"]["content"])
        self.assertIn("skill_routing", context)

    def test_uncategorized_query_calls_llm_without_skill_context(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            llm = CapturingLlm()
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                ReportingKnowledgeBase(root / "reference", root / "knowledge", KnowledgeDatabase(root / "knowledge.db")),
                llm,
            )
            payload = service.handle_message(None, "Can you make this sound more concise?")

        self.assertNotIn("selected_skill", payload)
        context = llm.calls[0]["context_pack"]
        self.assertNotIn("selected_reporting_skill", context)
        self.assertNotIn("selected_skill_spec", context)
        self.assertNotIn("skill_routing", context)

    def test_uncategorized_query_does_not_retrieve_or_display_rule_context(self) -> None:
        rule = RuleChunk(
            chunk_id="rule_1",
            source="ASC",
            citation="ASC 230",
            title="ASC 230 - Cash Flow Statements",
            text="Statement of cash flows guidance.",
            path="/tmp/asc230.pdf",
            score=10,
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            llm = CapturingLlm()
            knowledge = StubKnowledge(rule_chunks=[rule])
            service = ReportingChatService(SessionStore(root / "sessions"), knowledge, llm)
            payload = service.handle_message(None, "Can you make this sentence more concise?")

        self.assertEqual(knowledge.retrieve_calls, [])
        self.assertEqual(payload["citations"], [])
        self.assertNotIn("support_status", payload["display"])
        self.assertEqual(llm.calls[0]["context_pack"]["retrieved_reporting_guidance"], [])

    def test_rule_retrieval_query_is_skill_aware(self) -> None:
        rule = RuleChunk(
            chunk_id="rule_1",
            source="ASC",
            citation="ASC 230",
            title="ASC 230 - Cash Flow Statements",
            text="Statement of cash flows guidance.",
            path="/tmp/asc230.pdf",
            score=10,
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            llm = CapturingLlm()
            knowledge = StubKnowledge(rule_chunks=[rule])
            service = ReportingChatService(SessionStore(root / "sessions"), knowledge, llm)
            service.handle_message(None, "Draft my SCF from this support workbook")

        query, limit = knowledge.retrieve_calls[0]
        self.assertEqual(limit, 5)
        self.assertIn("ASC 230", query)
        self.assertIn("SCF Generation Skill", query)

    def test_reporting_query_without_sources_returns_no_source_state_without_download(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                StubKnowledge(),
                CapturingLlm(),
            )
            payload = service.handle_message(None, "What does ASC 999 require?")

            self.assertEqual(payload["display"]["support_status"], "no_source_found")
            self.assertEqual(payload["display"]["support_label"], "No source found")
            self.assertIn("I did not find relevant", payload["answer"])
            self.assertNotIn("artifacts", payload)
            self.assertFalse(list((root / "sessions" / payload["session_id"] / "artifacts").glob("*")))

    def test_source_heavy_query_with_rules_only_returns_partial_state(self) -> None:
        rule = RuleChunk(
            chunk_id="rule_1",
            source="ASC",
            citation="ASC 230",
            title="ASC 230 - Cash Flow Statements",
            text="Statement of cash flows guidance.",
            path="/tmp/asc230.pdf",
            score=10,
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            service = ReportingChatService(
                SessionStore(root / "sessions"),
                StubKnowledge(rule_chunks=[rule]),
                CapturingLlm(),
            )
            payload = service.handle_message(None, "Draft SEC filing language for this cash flow disclosure")

            self.assertEqual(payload["display"]["support_status"], "partially_supported")
            self.assertEqual(payload["display"]["support_label"], "Partially supported")
            self.assertEqual(payload["citations"][0]["citation"], "ASC 230")
            self.assertEqual(payload["citations"][0]["source_type"], "ASC")
            self.assertIn("excerpt", payload["citations"][0])
            self.assertEqual(payload["display"]["source_citations"][0]["number"], 1)
            self.assertNotIn("artifacts", payload)

    def test_global_rule_ingestion_uses_separate_tables(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            raw = root / "raw"
            asc = raw / "FASB Accounting Standard Codifications (ASC)"
            sec = raw / "SEC Regulation"
            asc.mkdir(parents=True)
            sec.mkdir(parents=True)
            (asc / "ASC 230 - Cash Flow Statements.pdf").write_bytes(b"not a real pdf")
            (sec / "SEC Regulation S-X.pdf").write_bytes(b"not a real pdf")

            db = KnowledgeDatabase(root / "knowledge.db")
            result = db.ingest_global_rules(raw)

            self.assertEqual(result["documents"], 2)
            self.assertGreaterEqual(result["chunks"], 2)
            self.assertGreater(len(db.retrieve_global("cash flow statement", limit=3)), 0)

    def test_response_formatter_removes_markdown_and_structures_answer(self) -> None:
        display = format_chat_response(
            "**ASC 230** is the primary guidance for cash flow presentation. "
            "It requires reporting cash receipts and payments during the period. "
            "- Use operating, investing, and financing categories. "
            "Next, tie the generated amounts to source support.",
            [{"citation": "ASC 230", "title": "ASC 230 - Cash Flow Statements"}],
        )

        self.assertNotIn("**", display["summary"])
        self.assertIn("ASC 230", display["summary"])
        self.assertLessEqual(len(display["key_points"]), 4)
        self.assertEqual(display["rule_support"][0]["citation"], "ASC 230")
        self.assertEqual(display["source_citations"][0]["source_type"], "SOURCE")

    def test_response_formatter_polishes_visible_sentences(self) -> None:
        display = format_chat_response(
            "answer starts lowercase and has no final punctuation",
            [],
        )

        self.assertEqual(display["summary"], "Answer starts lowercase and has no final punctuation.")

    def test_response_formatter_keeps_dash_spacing_grammatical(self) -> None:
        display = format_chat_response(
            "Yes\u2014please paste the sentence.",
            [],
        )

        self.assertEqual(display["summary"], "Yes, please paste the sentence.")


if __name__ == "__main__":
    unittest.main()
