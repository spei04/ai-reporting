from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl
from ai_reporting.engine import CashFlowEngine
from ai_reporting.mapping_review import build_mapping_review
from ai_reporting.parser_profile import MappingOverride, ParserProfile
from ai_reporting.profile import build_workbook_profile
from ai_reporting.source_map import SourceCellMap, SourceCellSignature
from ai_reporting.template import CashFlowTemplate, TemplateCell, TemplateSheet
from ai_reporting.validation import compare_to_answer
from ai_reporting.workbook import SourceWorkbook
import ai_reporting.engine as engine_module
import ai_reporting.workbook as workbook_module


ROOT = Path(__file__).resolve().parents[1]
REFERENCE = Path("/Users/serenapei/Downloads/ai_reporting/reference")
INPUT = REFERENCE / "FS-4 26Q1 SCF - 3rd Consol.xlsx"
ANSWER = REFERENCE / "answer.xlsx"
TEMPLATE = ROOT / "config" / "standard_q1_scf_template.json"
SOURCE_MAP = ROOT / "config" / "source_cell_map_q1.json"


class FirstMilestoneTest(unittest.TestCase):
    def test_generation_does_not_read_answer_workbook(self) -> None:
        engine = CashFlowEngine(INPUT, CashFlowTemplate.load(TEMPLATE), SourceCellMap.load(SOURCE_MAP))

        original_workbook_loader = workbook_module.openpyxl.load_workbook

        def guarded_load_workbook(path, *args, **kwargs):
            if Path(path).resolve() == ANSWER.resolve():
                raise AssertionError("Generation attempted to read answer.xlsx")
            return original_workbook_loader(path, *args, **kwargs)

        with tempfile.TemporaryDirectory() as directory, patch.object(
            workbook_module.openpyxl, "load_workbook", guarded_load_workbook
        ), patch.object(engine_module.openpyxl, "load_workbook", guarded_load_workbook):
            result = engine.generate(Path(directory))

            self.assertTrue(result.output_workbook.exists())
            self.assertTrue(result.values_json.exists())
            self.assertTrue(result.evidence_json.exists())
            self.assertTrue(result.normalized_support_json.exists())
            self.assertTrue(result.workbook_profile_json.exists())
            self.assertTrue(result.mapping_review_json.exists())

    def test_q1_detailed_bridge_matches_second_answer_tab(self) -> None:
        engine = CashFlowEngine(INPUT, CashFlowTemplate.load(TEMPLATE), SourceCellMap.load(SOURCE_MAP))
        with tempfile.TemporaryDirectory() as directory:
            result = engine.generate(Path(directory))

            self.assertTrue(result.output_workbook.exists())
            self.assertTrue(result.values_json.exists())
            self.assertTrue(result.evidence_json.exists())
            self.assertTrue(result.normalized_support_json.exists())
            self.assertTrue(result.workbook_profile_json.exists())
            self.assertTrue(result.mapping_review_json.exists())

            diffs = compare_to_answer(engine, ANSWER)
            self.assertEqual(diffs, [])

            generated = openpyxl.load_workbook(result.output_workbook, data_only=False)
            self.assertIn("Evidence Index", generated.sheetnames)
            evidence_index = generated["Evidence Index"]
            self.assertEqual(evidence_index["A1"].value, "Output Sheet")
            evidence_rows = {
                f"{row[0]}!{row[1]}": row_number
                for row_number, row in enumerate(
                    evidence_index.iter_rows(min_row=2, values_only=True),
                    start=2,
                )
            }
            self.assertIn("2. 26Q1 QTD!B25", evidence_rows)

            output_cell = generated["2. 26Q1 QTD"]["B25"]
            self.assertIsNotNone(output_cell.hyperlink)
            self.assertEqual(output_cell.hyperlink.target, f"#'Evidence Index'!A{evidence_rows['2. 26Q1 QTD!B25']}")
            self.assertIsNotNone(output_cell.comment)
            self.assertIn("Evidence Index row", output_cell.comment.text)
            self.assertIn("ASC 230", output_cell.comment.text)

            evidence = json.loads(result.evidence_json.read_text())
            traced = {
                f"{item['output_sheet']}!{item['output_cell']}": item
                for item in evidence
            }
            self.assertIn("2. 26Q1 QTD!B25", traced)
            item = traced["2. 26Q1 QTD!B25"]
            self.assertIn("output_value", item)
            self.assertIn("output_formula", item)
            self.assertEqual(item["review_status"], "Linked")
            self.assertEqual(item["rule_reference"], "ASC 230")
            self.assertGreater(len(item["dependency_details"]), 0)
            self.assertGreater(len(item["source_locations"]), 0)

    def test_profile_resolves_slightly_different_sheet_names(self) -> None:
        profile = build_workbook_profile(
            [
                "Balance Sheet",
                "Income Statement",
                "Short-term Investments",
                "Mark to Market",
                "Hedge Rollforward",
                "Lease Support",
                "Property and Equipment",
                "Unpaid CapEx",
                "Capitalized Stock Comp",
                "Purchase Price Allocation",
                "Equity Rollforward",
                "SBC Expense",
                "Share Repurchase",
                "CECL Reserves",
                "Sale of Strategic Investment",
            ]
        )

        self.assertEqual(profile.actual_sheet("BS"), "Balance Sheet")
        self.assertEqual(profile.actual_sheet("IS"), "Income Statement")
        self.assertEqual(profile.actual_sheet("3. STI"), "Short-term Investments")
        self.assertEqual(profile.actual_sheet("6. Fixed Assets"), "Property and Equipment")
        self.assertEqual(profile.actual_sheet("8. Equity RF"), "Equity Rollforward")
        self.assertEqual(profile.actual_sheet("10. CECL_E&O"), "CECL Reserves")

    def test_source_map_resolves_moved_formula_cell(self) -> None:
        formula = "=H9+H20+H31"
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "moved_formula.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "3. STI"
            ws["Q3"] = "old location"
            ws["R4"] = formula
            wb.save(path)

            source_map = SourceCellMap(
                {
                    "3. STI!Q3": SourceCellSignature(
                        canonical_sheet="3. STI",
                        canonical_cell="Q3",
                        value_type="NoneType",
                        formula=formula,
                        row_label=None,
                        column_label=None,
                    )
                }
            )
            workbook = SourceWorkbook(path, source_map)

            self.assertEqual(workbook.actual_cell_name("3. STI", "Q3"), "R4")
            self.assertEqual(workbook.formula("3. STI", "Q3"), formula)

    def test_evidence_json_includes_inspector_fields(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "support.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BS"
            ws["A1"] = 100
            ws["A2"] = 25
            wb.save(source_path)

            template = CashFlowTemplate(
                "test",
                {
                    "1. SCF": TemplateSheet(
                        "1. SCF",
                        {
                            "B2": TemplateCell("B2", formula="='BS'!A1+'BS'!A2"),
                        },
                    ),
                    "2. 26Q1 QTD": TemplateSheet("2. 26Q1 QTD", {}),
                    "2a. 2026 YTD": TemplateSheet("2a. 2026 YTD", {}),
                },
            )
            engine = CashFlowEngine(source_path, template)
            result = engine.generate(root / "out")

            evidence = json.loads(result.evidence_json.read_text())
            self.assertEqual(len(evidence), 1)
            item = evidence[0]
            self.assertEqual(item["output_sheet"], "1. SCF")
            self.assertEqual(item["output_cell"], "B2")
            self.assertEqual(item["output_value"], 125.0)
            self.assertEqual(item["output_formula"], "='BS'!A1+'BS'!A2")
            self.assertEqual(item["rule_reference"], "ASC 230")
            self.assertEqual(item["review_status"], "Linked")
            self.assertEqual(item["source_locations"], ["BS!A1", "BS!A2"])
            self.assertEqual(len(item["dependency_details"]), 2)

    def test_evidence_links_include_transitive_generated_sources(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "support.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Support"
            ws["A1"] = 100
            wb.save(source_path)

            template = CashFlowTemplate(
                "test",
                {
                    "1. SCF": TemplateSheet("1. SCF", {}),
                    "2. 26Q1 QTD": TemplateSheet(
                        "2. 26Q1 QTD",
                        {
                            "B2": TemplateCell("B2", formula="=-SUM(C2:D2)"),
                            "C2": TemplateCell("C2", formula="='Support'!A1"),
                        },
                    ),
                    "2a. 2026 YTD": TemplateSheet("2a. 2026 YTD", {}),
                },
            )
            engine = CashFlowEngine(source_path, template)
            result = engine.generate(root / "out")

            evidence = json.loads(result.evidence_json.read_text())
            traced = {
                f"{item['output_sheet']}!{item['output_cell']}": item
                for item in evidence
            }

            item = traced["2. 26Q1 QTD!B2"]
            self.assertEqual(item["review_status"], "Linked")
            self.assertEqual(item["source_locations"], ["Support!A1"])

    def test_source_map_resolves_moved_labeled_value_cell(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "moved_value.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Support"
            ws["C1"] = "Q1 2026"
            ws["B3"] = "Net income"
            ws["C3"] = 123
            wb.save(path)

            source_map = SourceCellMap(
                {
                    "Support!B2": SourceCellSignature(
                        canonical_sheet="Support",
                        canonical_cell="B2",
                        value_type="int",
                        formula=None,
                        row_label="net income",
                        column_label="q1 2026",
                    )
                }
            )
            workbook = SourceWorkbook(path, source_map)

            self.assertEqual(workbook.actual_cell_name("Support", "B2"), "C3")
            self.assertEqual(workbook.value("Support", "B2"), 123)

    def test_mapping_review_reports_exact_moved_and_missing_items(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "mapping_review.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Support"
            ws["A1"] = "Exact label"
            ws["B1"] = 10
            ws["D1"] = "Q1 2026"
            ws["C4"] = "Moved label"
            ws["D4"] = 25
            wb.save(path)

            source_map = SourceCellMap(
                {
                    "Support!B1": SourceCellSignature(
                        canonical_sheet="Support",
                        canonical_cell="B1",
                        value_type="int",
                        formula=None,
                        row_label="exact label",
                        column_label=None,
                    ),
                    "Support!B3": SourceCellSignature(
                        canonical_sheet="Support",
                        canonical_cell="B3",
                        value_type="int",
                        formula=None,
                        row_label="moved label",
                        column_label="q1 2026",
                    ),
                    "Missing!A1": SourceCellSignature(
                        canonical_sheet="Missing",
                        canonical_cell="A1",
                        value_type="int",
                        formula=None,
                        row_label=None,
                        column_label=None,
                    ),
                }
            )

            review = build_mapping_review(SourceWorkbook(path, source_map))
            by_key = {f"{item.canonical_sheet}!{item.canonical_cell}": item for item in review}

            self.assertEqual(by_key["Support!B1"].status, "exact")
            self.assertEqual(by_key["Support!B3"].status, "moved")
            self.assertEqual(by_key["Support!B3"].actual_cell, "D4")
            self.assertEqual(by_key["Missing!A1"].status, "missing_sheet")

    def test_parser_profile_override_takes_priority_over_learned_match(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "override.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Support"
            ws["A1"] = "Expected label"
            ws["B1"] = 10
            ws["D4"] = 42
            wb.save(path)

            source_map = SourceCellMap(
                {
                    "Support!B1": SourceCellSignature(
                        canonical_sheet="Support",
                        canonical_cell="B1",
                        value_type="int",
                        formula=None,
                        row_label="expected label",
                        column_label=None,
                    )
                }
            )
            parser_profile = ParserProfile(
                {
                    "Support!B1": MappingOverride(
                        canonical_sheet="Support",
                        canonical_cell="B1",
                        actual_sheet="Support",
                        actual_cell="D4",
                        note="Reviewer selected alternate support cell.",
                    )
                }
            )
            workbook = SourceWorkbook(path, source_map, parser_profile)

            self.assertEqual(workbook.actual_cell_name("Support", "B1"), "D4")
            self.assertEqual(workbook.cell_resolution_status("Support", "B1"), "manual_override")
            self.assertEqual(workbook.value("Support", "B1"), 42)

            review = build_mapping_review(workbook)
            self.assertEqual(review[0].status, "manual_override")
            self.assertEqual(review[0].actual_cell, "D4")

    def test_parser_profile_save_and_load_roundtrip(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "company_parser_profile.json"
            profile = ParserProfile().with_override(
                MappingOverride(
                    canonical_sheet="Support",
                    canonical_cell="B1",
                    actual_sheet="Uploaded Support",
                    actual_cell="D4",
                    note="Reviewer approved.",
                )
            )

            profile.save(path)
            loaded = ParserProfile.load(path)
            override = loaded.override_for("Support", "B1")

            self.assertIsNotNone(override)
            self.assertEqual(override.actual_sheet, "Uploaded Support")
            self.assertEqual(override.actual_cell, "D4")
            self.assertEqual(override.note, "Reviewer approved.")


if __name__ == "__main__":
    unittest.main()
