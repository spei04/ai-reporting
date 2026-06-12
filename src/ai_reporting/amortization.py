from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class AmortizationTerms:
    principal: float | None = None
    annual_rate: float | None = None
    years: int | None = None
    payments_per_year: int = 12

    @property
    def payment_count(self) -> int:
        return int((self.years or 0) * self.payments_per_year)

    @property
    def monthly_rate(self) -> float:
        return (self.annual_rate or 0.0) / self.payments_per_year


def is_amortization_request(message: str) -> bool:
    text = message.lower()
    return "amortization" in text and ("schedule" in text or "generate" in text or "create" in text)


def parse_amortization_terms(message: str, recent_messages: list[dict[str, Any]] | None = None) -> AmortizationTerms:
    combined = _combined_text(message, recent_messages or [])
    return AmortizationTerms(
        principal=_parse_principal(combined),
        annual_rate=_parse_rate(combined),
        years=_parse_years(combined) or 3,
    )


def missing_amortization_inputs(terms: AmortizationTerms) -> list[str]:
    missing = []
    if terms.principal is None:
        missing.append("principal amount")
    if terms.annual_rate is None:
        missing.append("fixed annual interest rate")
    return missing


def generate_amortization_workbook(terms: AmortizationTerms, output_path: Path) -> dict[str, Any]:
    if terms.principal is None or terms.annual_rate is None or terms.years is None:
        raise ValueError("Principal, annual rate, and term are required")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    schedule = workbook.create_sheet("Amortization Schedule")

    principal = Decimal(str(terms.principal))
    periodic_rate = Decimal(str(terms.monthly_rate))
    payment_count = terms.payment_count
    payment = _payment(principal, periodic_rate, payment_count)

    summary_rows = [
        ("Principal", float(principal)),
        ("Annual fixed interest rate", terms.annual_rate),
        ("Term in years", terms.years),
        ("Payment frequency", "Monthly"),
        ("Number of payments", payment_count),
        ("Scheduled payment", float(payment)),
    ]
    summary["A1"] = "Amortization Schedule Inputs"
    summary["A1"].font = Font(bold=True, size=14)
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        summary.cell(row=row_index, column=1, value=label)
        summary.cell(row=row_index, column=2, value=value)
    summary["B4"].number_format = "0.00%"
    summary["B3"].number_format = "$#,##0.00"
    summary["B8"].number_format = "$#,##0.00"
    _autosize(summary)

    headers = ["Period", "Beginning Balance", "Payment", "Interest Expense", "Principal Payment", "Ending Balance"]
    for column, header in enumerate(headers, start=1):
        cell = schedule.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E6F73")

    balance = principal
    total_interest = Decimal("0")
    for period in range(1, payment_count + 1):
        interest = _money(balance * periodic_rate)
        principal_payment = min(_money(payment - interest), balance)
        actual_payment = _money(interest + principal_payment)
        ending_balance = _money(balance - principal_payment)
        if period == payment_count:
            principal_payment = balance
            actual_payment = _money(interest + principal_payment)
            ending_balance = Decimal("0.00")

        row = [period, balance, actual_payment, interest, principal_payment, ending_balance]
        for column, value in enumerate(row, start=1):
            cell = schedule.cell(row=period + 1, column=column, value=float(value) if isinstance(value, Decimal) else value)
            if column > 1:
                cell.number_format = "$#,##0.00"

        total_interest += interest
        balance = ending_balance

    total_row = payment_count + 3
    schedule.cell(row=total_row, column=1, value="Totals").font = Font(bold=True)
    schedule.cell(row=total_row, column=3, value=f"=SUM(C2:C{payment_count + 1})")
    schedule.cell(row=total_row, column=4, value=f"=SUM(D2:D{payment_count + 1})")
    schedule.cell(row=total_row, column=5, value=f"=SUM(E2:E{payment_count + 1})")
    for column in (3, 4, 5):
        schedule.cell(row=total_row, column=column).number_format = "$#,##0.00"
        schedule.cell(row=total_row, column=column).font = Font(bold=True)

    _autosize(schedule)
    workbook.save(output_path)
    return {
        "principal": float(principal),
        "annual_rate": terms.annual_rate,
        "years": terms.years,
        "payments_per_year": terms.payments_per_year,
        "payment_count": payment_count,
        "scheduled_payment": float(payment),
        "total_interest": float(_money(total_interest)),
    }


def _combined_text(message: str, recent_messages: list[dict[str, Any]]) -> str:
    parts = [str(item.get("content", "")) for item in recent_messages[-6:] if item.get("role") == "user"]
    parts.append(message)
    return "\n".join(parts)


def _parse_principal(text: str) -> float | None:
    patterns = [
        r"(?:principal|principle|loan amount|amount)\s*(?:is|=|:)?\s*\$?\s*([0-9][0-9,]*(?:\.\d+)?)",
        r"\$\s*([0-9][0-9,]*(?:\.\d+)?)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return float(match.group(1).replace(",", ""))
    return None


def _parse_rate(text: str) -> float | None:
    patterns = [
        r"(?:interest rate|annual rate|rate|apr)\s*(?:is|=|:)?\s*([0-9]+(?:\.\d+)?)\s*%",
        r"([0-9]+(?:\.\d+)?)\s*%\s*(?:interest|annual|fixed|rate|apr)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return float(match.group(1)) / 100
    return None


def _parse_years(text: str) -> int | None:
    match = re.search(r"(?:over|for|term|in)\s*(\d+)\s*years?", text, flags=re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.search(r"(\d+)\s*years?", text, flags=re.IGNORECASE)
    if match:
        return int(match.group(1))
    match = re.search(r"(\d+)\s*months?", text, flags=re.IGNORECASE)
    if match:
        months = int(match.group(1))
        return max(1, round(months / 12))
    return None


def _payment(principal: Decimal, periodic_rate: Decimal, periods: int) -> Decimal:
    if periods <= 0:
        raise ValueError("Payment count must be greater than zero")
    if periodic_rate == 0:
        return _money(principal / Decimal(periods))
    factor = (Decimal("1") + periodic_rate) ** periods
    return _money(principal * periodic_rate * factor / (factor - Decimal("1")))


def _money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _autosize(sheet) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 28)
