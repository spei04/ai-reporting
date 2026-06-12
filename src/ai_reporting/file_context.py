from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import openpyxl


def summarize_upload(filename: str, data: bytes) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return summarize_xlsx(data)
    if suffix == ".csv":
        return summarize_csv(data)
    if suffix in {".txt", ".md"}:
        return summarize_text(data)
    if suffix == ".pdf":
        return summarize_pdf(data)
    return {"type": suffix.lstrip(".") or "unknown", "note": "Stored without structured extraction."}


def summarize_xlsx(data: bytes) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(BytesIO(data), data_only=False, read_only=True)
    sheets = []
    for ws in workbook.worksheets[:20]:
        non_empty = 0
        formulas = 0
        samples = []
        for row in ws.iter_rows(max_row=min(ws.max_row, 25), max_col=min(ws.max_column, 12)):
            values = []
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty += 1
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas += 1
                    values.append(str(cell.value)[:80])
            if values and len(samples) < 8:
                samples.append(values)
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "sample_non_empty_cells": non_empty,
                "sample_formula_cells": formulas,
                "sample_rows": samples,
            }
        )
    return {"type": "xlsx", "sheet_count": len(workbook.sheetnames), "sheets": sheets}


def summarize_csv(data: bytes) -> dict[str, Any]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = []
    for index, row in enumerate(reader):
        if index >= 25:
            break
        rows.append(row[:20])
    return {
        "type": "csv",
        "headers": rows[0] if rows else [],
        "sample_rows": rows[1:8],
        "sample_row_count": len(rows),
    }


def summarize_text(data: bytes) -> dict[str, Any]:
    text = data.decode("utf-8", errors="replace")
    return {"type": "text", "preview": text[:4000], "character_count": len(text)}


def summarize_pdf(data: bytes) -> dict[str, Any]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(data))
        text_parts = []
        for page in reader.pages[:5]:
            text_parts.append((page.extract_text() or "")[:1200])
        return {
            "type": "pdf",
            "page_count": len(reader.pages),
            "preview": "\n".join(text_parts)[:4000],
        }
    except Exception as exc:
        return {"type": "pdf", "note": f"Stored PDF; text extraction failed: {exc}"}


def extract_text_chunks(filename: str, data: bytes, summary: dict[str, Any], size: int = 1600) -> list[str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        text = workbook_summary_text(summary)
    elif suffix == ".csv":
        text = csv_summary_text(summary)
    elif suffix in {".txt", ".md"}:
        text = data.decode("utf-8", errors="replace")
    elif suffix == ".pdf":
        text = pdf_text(data) or str(summary.get("preview") or summary)
    else:
        text = str(summary)
    return chunk_text(text, size=size)


def workbook_summary_text(summary: dict[str, Any]) -> str:
    lines = [f"Workbook with {summary.get('sheet_count', 0)} sheets."]
    for sheet in summary.get("sheets", []):
        lines.append(f"Sheet: {sheet.get('name')} rows={sheet.get('max_row')} columns={sheet.get('max_column')}")
        for row in sheet.get("sample_rows", []):
            lines.append(" | ".join(str(value) for value in row))
    return "\n".join(lines)


def csv_summary_text(summary: dict[str, Any]) -> str:
    lines = ["CSV upload", "Headers: " + " | ".join(str(item) for item in summary.get("headers", []))]
    for row in summary.get("sample_rows", []):
        lines.append(" | ".join(str(value) for value in row))
    return "\n".join(lines)


def pdf_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader

        reader = PdfReader(BytesIO(data))
        return "\n".join((page.extract_text() or "") for page in reader.pages[:20])
    except Exception:
        return ""


def chunk_text(text: str, size: int = 1600, overlap: int = 200) -> list[str]:
    cleaned = " ".join(str(text or "").split())
    if not cleaned:
        return ["No extractable text."]
    chunks = []
    start = 0
    while start < len(cleaned):
        chunks.append(cleaned[start : start + size])
        start += size - overlap
    return chunks
