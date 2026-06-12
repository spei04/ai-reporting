from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter

from .parser_profile import ParserProfile
from .profile import build_workbook_profile
from .source_map import SourceCellMap, SourceCellSignature


@dataclass(frozen=True)
class SheetSummary:
    title: str
    max_row: int
    max_column: int


@dataclass(frozen=True)
class WorkbookSummary:
    path: Path
    sheet_count: int
    sheets: list[SheetSummary]
    profile: dict[str, object]


class SourceWorkbook:
    """Read-only wrapper around the uploaded support workbook."""

    def __init__(
        self,
        path: Path,
        source_map: SourceCellMap | None = None,
        parser_profile: ParserProfile | None = None,
    ):
        self.path = path
        self.values = openpyxl.load_workbook(path, data_only=True, read_only=True)
        self.formulas = openpyxl.load_workbook(path, data_only=False, read_only=True)
        self.profile = build_workbook_profile(self.formulas.sheetnames)
        self.source_map = source_map
        self.parser_profile = parser_profile or ParserProfile()
        self._resolved_cells: dict[tuple[str, str], str] = {}
        self._resolved_sheets: dict[tuple[str, str], str | None] = {}
        self._resolution_status: dict[tuple[str, str], str] = {}

    def summary(self) -> WorkbookSummary:
        return WorkbookSummary(
            path=self.path,
            sheet_count=len(self.formulas.worksheets),
            sheets=[
                SheetSummary(ws.title, ws.max_row, ws.max_column)
                for ws in self.formulas.worksheets
            ],
            profile=self.profile.to_json(),
        )

    def value(self, sheet_name: str, cell: str) -> Any:
        actual_sheet = self.resolve_cell_sheet(sheet_name, cell)
        if actual_sheet is None:
            return None
        actual_cell = self.resolve_cell(sheet_name, cell)
        return self.values[actual_sheet][actual_cell].value

    def formula(self, sheet_name: str, cell: str) -> Any:
        actual_sheet = self.resolve_cell_sheet(sheet_name, cell)
        if actual_sheet is None:
            return None
        actual_cell = self.resolve_cell(sheet_name, cell)
        return self.formulas[actual_sheet][actual_cell].value

    def resolve_sheet(self, sheet_name: str) -> str | None:
        if sheet_name in self.values.sheetnames:
            return sheet_name
        return self.profile.actual_sheet(sheet_name)

    def has_sheet(self, sheet_name: str) -> bool:
        return self.resolve_sheet(sheet_name) is not None

    def value_sheet(self, sheet_name: str):
        actual_sheet = self.resolve_sheet(sheet_name)
        if actual_sheet is None:
            return None
        return self.values[actual_sheet]

    def actual_sheet_name(self, sheet_name: str) -> str | None:
        return self.resolve_sheet(sheet_name)

    def resolve_cell(self, sheet_name: str, cell: str) -> str:
        cache_key = (sheet_name, cell)
        if cache_key in self._resolved_cells:
            return self._resolved_cells[cache_key]

        override = self.parser_profile.override_for(sheet_name, cell)
        if override is not None:
            actual_sheet = self.resolve_sheet(override.actual_sheet)
            self._resolved_sheets[cache_key] = actual_sheet
            self._resolved_cells[cache_key] = override.actual_cell
            self._resolution_status[cache_key] = "manual_override" if actual_sheet else "missing_sheet"
            return override.actual_cell

        actual_sheet = self.resolve_sheet(sheet_name)
        self._resolved_sheets[cache_key] = actual_sheet
        if actual_sheet is None:
            self._resolution_status[cache_key] = "missing_sheet"
            return cell

        if self.source_map is None:
            self._resolution_status[cache_key] = "unmapped"
            return cell

        signature = self.source_map.signature(sheet_name, cell)
        if signature is None:
            self._resolution_status[cache_key] = "unmapped"
            return cell

        exact_formula = self.formulas[actual_sheet][cell].value
        exact_value = self.values[actual_sheet][cell].value
        if _cell_matches_signature(
            signature,
            exact_value,
            exact_formula,
            _nearest_left_label(self.values[actual_sheet], cell),
            _nearest_above_label(self.values[actual_sheet], cell),
        ):
            self._resolved_cells[cache_key] = cell
            self._resolution_status[cache_key] = "exact"
            return cell

        resolved = self._find_cell_by_signature(actual_sheet, signature)
        self._resolved_cells[cache_key] = resolved or cell
        self._resolution_status[cache_key] = "moved" if resolved else "needs_review"
        return self._resolved_cells[cache_key]

    def _find_cell_by_signature(self, actual_sheet: str, signature: SourceCellSignature) -> str | None:
        ws_formula = self.formulas[actual_sheet]
        ws_value = self.values[actual_sheet]

        if signature.formula:
            for row in range(1, ws_formula.max_row + 1):
                for column in range(1, ws_formula.max_column + 1):
                    if ws_formula.cell(row, column).value == signature.formula:
                        return f"{get_column_letter(column)}{row}"

        best: tuple[int, str] | None = None
        for row in range(1, ws_value.max_row + 1):
            for column in range(1, ws_value.max_column + 1):
                coordinate = f"{get_column_letter(column)}{row}"
                value = ws_value.cell(row, column).value
                score = _signature_score(
                    signature,
                    value,
                    self.formulas[actual_sheet][coordinate].value,
                    _nearest_left_label(ws_value, coordinate),
                    _nearest_above_label(ws_value, coordinate),
                )
                if score > 0 and (best is None or score > best[0]):
                    best = (score, coordinate)

        return best[1] if best else None

    def actual_cell_name(self, sheet_name: str, cell: str) -> str:
        return self.resolve_cell(sheet_name, cell)

    def resolve_cell_sheet(self, sheet_name: str, cell: str) -> str | None:
        cache_key = (sheet_name, cell)
        if cache_key not in self._resolved_sheets:
            self.resolve_cell(sheet_name, cell)
        return self._resolved_sheets.get(cache_key, self.resolve_sheet(sheet_name))

    def cell_resolution_status(self, sheet_name: str, cell: str) -> str:
        cache_key = (sheet_name, cell)
        if cache_key not in self._resolution_status:
            self.resolve_cell(sheet_name, cell)
        return self._resolution_status.get(cache_key, "unmapped")


def _cell_matches_signature(
    signature: SourceCellSignature,
    value: Any,
    formula: Any,
    row_label: str | None,
    column_label: str | None,
) -> bool:
    if signature.formula:
        return formula == signature.formula
    if _value_type(value) != signature.value_type:
        return False
    if signature.row_label and signature.row_label != row_label:
        return False
    if signature.column_label and signature.column_label != column_label:
        return False
    return True


def _signature_score(
    signature: SourceCellSignature,
    value: Any,
    formula: Any,
    row_label: str | None,
    column_label: str | None,
) -> int:
    score = 0
    if signature.formula and formula == signature.formula:
        score += 100
    if _value_type(value) == signature.value_type:
        score += 1
    if signature.row_label and signature.row_label == row_label:
        score += 10
    if signature.column_label and signature.column_label == column_label:
        score += 10
    return score if score >= 10 else 0


def _nearest_left_label(ws, cell: str) -> str | None:
    row, col = coordinate_to_tuple(cell)
    for candidate_col in range(col - 1, max(col - 8, 0), -1):
        value = ws.cell(row, candidate_col).value
        if isinstance(value, str) and value.strip():
            return _normalize_label(value)
    return None


def _nearest_above_label(ws, cell: str) -> str | None:
    row, col = coordinate_to_tuple(cell)
    for candidate_row in range(row - 1, max(row - 8, 0), -1):
        value = ws.cell(candidate_row, col).value
        if isinstance(value, str) and value.strip():
            return _normalize_label(value)
    return None


def _normalize_label(value: str) -> str:
    import re

    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _value_type(value: Any) -> str:
    return type(value).__name__
