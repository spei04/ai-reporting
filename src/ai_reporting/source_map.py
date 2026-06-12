from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple

from .formula import SHEET_REF_RE


GENERATED_OUTPUT_SHEETS = {"1. SCF", "2. 26Q1 QTD", "2a. 2026 YTD"}


@dataclass(frozen=True)
class SourceCellSignature:
    canonical_sheet: str
    canonical_cell: str
    value_type: str
    formula: str | None
    row_label: str | None
    column_label: str | None


class SourceCellMap:
    def __init__(self, signatures: dict[str, SourceCellSignature]):
        self.signatures = signatures

    @classmethod
    def load(cls, path: Path | None) -> "SourceCellMap | None":
        if path is None or not path.exists():
            return None
        data = json.loads(path.read_text())
        signatures = {
            item["key"]: SourceCellSignature(
                canonical_sheet=item["canonical_sheet"],
                canonical_cell=item["canonical_cell"],
                value_type=item["value_type"],
                formula=item.get("formula"),
                row_label=item.get("row_label"),
                column_label=item.get("column_label"),
            )
            for item in data["source_cells"]
        }
        return cls(signatures)

    def signature(self, canonical_sheet: str, canonical_cell: str) -> SourceCellSignature | None:
        return self.signatures.get(_key(canonical_sheet, canonical_cell))

    def to_json(self) -> dict[str, Any]:
        return {
            "source_cells": [
                {"key": key, **asdict(signature)}
                for key, signature in sorted(self.signatures.items())
            ]
        }


def build_source_cell_map(
    template_path: Path,
    approved_workbook_path: Path,
    output_path: Path,
) -> SourceCellMap:
    source_refs = _extract_source_refs(template_path)
    formulas = openpyxl.load_workbook(approved_workbook_path, data_only=False, read_only=True)
    values = openpyxl.load_workbook(approved_workbook_path, data_only=True, read_only=True)

    signatures: dict[str, SourceCellSignature] = {}
    for sheet_name, cell in sorted(source_refs):
        if sheet_name not in formulas.sheetnames:
            continue
        ws_formula = formulas[sheet_name]
        ws_value = values[sheet_name]
        formula = ws_formula[cell].value
        value = ws_value[cell].value
        signatures[_key(sheet_name, cell)] = SourceCellSignature(
            canonical_sheet=sheet_name,
            canonical_cell=cell,
            value_type=type(value).__name__,
            formula=formula if isinstance(formula, str) and formula.startswith("=") else None,
            row_label=_nearest_left_label(ws_value, cell),
            column_label=_nearest_above_label(ws_value, cell),
        )

    source_map = SourceCellMap(signatures)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(source_map.to_json(), indent=2, default=str))
    return source_map


def _extract_source_refs(template_path: Path) -> set[tuple[str, str]]:
    data = json.loads(template_path.read_text())
    refs: set[tuple[str, str]] = set()
    for sheet_data in data["sheets"].values():
        for item in sheet_data["cells"]:
            formula = item.get("formula")
            if not formula:
                continue
            for match in SHEET_REF_RE.finditer(formula):
                sheet = match.group("quoted") or match.group("plain")
                cell = f"{match.group('col')}{match.group('row')}"
                if sheet not in GENERATED_OUTPUT_SHEETS:
                    refs.add((sheet, cell))
    return refs


def _nearest_left_label(ws, cell: str) -> str | None:
    row, column = coordinate_to_tuple(cell)
    for col in range(column - 1, max(column - 8, 0), -1):
        value = ws.cell(row, col).value
        if _is_label(value):
            return _normalize_label(str(value))
    return None


def _nearest_above_label(ws, cell: str) -> str | None:
    row, column = coordinate_to_tuple(cell)
    for candidate_row in range(row - 1, max(row - 8, 0), -1):
        value = ws.cell(candidate_row, column).value
        if _is_label(value):
            return _normalize_label(str(value))
    return None


def _is_label(value: Any) -> bool:
    return isinstance(value, str) and bool(value.strip()) and not value.startswith("=")


def _normalize_label(value: str) -> str:
    value = value.lower().replace("&", "and")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _key(sheet_name: str, cell: str) -> str:
    return f"{sheet_name}!{cell}"
