from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font

from .formula import FormulaEvaluator
from .mapping_review import build_mapping_review, mapping_review_to_json
from .normalize import normalize_support, normalized_support_to_json
from .parser_profile import ParserProfile
from .source_map import SourceCellMap
from .template import CashFlowTemplate, TemplateCell
from .workbook import SourceWorkbook


GENERATED_SHEETS = ["1. SCF", "2. 26Q1 QTD", "2a. 2026 YTD"]
EVIDENCE_INDEX_SHEET = "Evidence Index"


@dataclass(frozen=True)
class EvidenceLink:
    output_sheet: str
    output_cell: str
    output_value: Any
    output_formula: str | None
    dependencies: list[str]
    dependency_details: list[dict[str, Any]]
    source_locations: list[str]
    rule_reference: str
    review_status: str


@dataclass(frozen=True)
class GenerationResult:
    output_workbook: Path
    values_json: Path
    evidence_json: Path
    normalized_support_json: Path
    workbook_profile_json: Path
    mapping_review_json: Path


class CashFlowEngine:
    def __init__(
        self,
        source_path: Path,
        template: CashFlowTemplate,
        source_map: SourceCellMap | None = None,
        parser_profile: ParserProfile | None = None,
    ):
        self.source = SourceWorkbook(source_path, source_map, parser_profile)
        self.template = template
        self._evaluating: set[tuple[str, str]] = set()
        self._value_cache: dict[tuple[str, str], Any] = {}
        self.evaluator = FormulaEvaluator(self._get_value)

    def generate(self, output_dir: Path) -> GenerationResult:
        output_dir.mkdir(parents=True, exist_ok=True)
        workbook_path = output_dir / "generated_scf_q1_2026.xlsx"
        values_path = output_dir / "generated_values.json"
        evidence_path = output_dir / "evidence_links.json"
        normalized_path = output_dir / "normalized_support.json"
        profile_path = output_dir / "workbook_profile.json"
        mapping_review_path = output_dir / "mapping_review.json"

        normalized = normalize_support(self.source)
        normalized_path.write_text(json.dumps(normalized_support_to_json(normalized), indent=2, default=str))
        profile_path.write_text(json.dumps(self.source.profile.to_json(), indent=2, default=str))
        mapping_review = build_mapping_review(self.source)
        mapping_review_path.write_text(
            json.dumps(mapping_review_to_json(mapping_review), indent=2, default=str)
        )

        generated_values = self._evaluate_all()
        values_path.write_text(json.dumps(self._json_values(generated_values), indent=2, default=str))

        evidence = self._evidence_links()
        evidence_path.write_text(json.dumps([asdict(item) for item in evidence], indent=2))

        self._write_workbook(workbook_path, evidence, generated_values)

        return GenerationResult(
            output_workbook=workbook_path,
            values_json=values_path,
            evidence_json=evidence_path,
            normalized_support_json=normalized_path,
            workbook_profile_json=profile_path,
            mapping_review_json=mapping_review_path,
        )

    def _template_cell(self, sheet_name: str, cell: str) -> TemplateCell | None:
        sheet = self.template.sheets.get(sheet_name)
        if not sheet:
            return None
        return sheet.cells.get(cell)

    def _get_value(self, sheet_name: str, cell: str) -> Any:
        key = (sheet_name, cell)
        if key in self._value_cache:
            return self._value_cache[key]

        template_cell = self._template_cell(sheet_name, cell)
        if template_cell is None:
            return self.source.value(sheet_name, cell)

        if key in self._evaluating:
            raise ValueError(f"Circular formula dependency detected at {sheet_name}!{cell}")

        self._evaluating.add(key)
        try:
            if template_cell.formula:
                value = self.evaluator.evaluate(sheet_name, cell, template_cell.formula)
            else:
                value = template_cell.value
            self._value_cache[key] = value
            return value
        finally:
            self._evaluating.remove(key)

    def _evaluate_all(self) -> dict[str, dict[str, Any]]:
        values: dict[str, dict[str, Any]] = {}
        for sheet_name in GENERATED_SHEETS:
            sheet = self.template.sheets.get(sheet_name)
            if not sheet:
                continue
            values[sheet_name] = {}
            for cell in sheet.cells:
                values[sheet_name][cell] = self._get_value(sheet_name, cell)
        return values

    def _json_values(self, values: dict[str, dict[str, Any]]) -> dict[str, Any]:
        return {
            "template_version": self.template.version,
            "generated_sheets": values,
        }

    def _evidence_links(self) -> list[EvidenceLink]:
        links: list[EvidenceLink] = []
        for output_key, dependencies in sorted(self.evaluator.dependencies.items()):
            sheet, cell = output_key.split("!", 1)
            if sheet in ("1. SCF", "2. 26Q1 QTD"):
                dependency_details = [
                    self._dependency_detail(dependency)
                    for dependency in sorted(dependencies)
                ]
                source_locations = self._source_locations_for_output(f"{sheet}!{cell}")
                template_cell = self._template_cell(sheet, cell)
                links.append(
                    EvidenceLink(
                        output_sheet=sheet,
                        output_cell=cell,
                        output_value=self._get_value(sheet, cell),
                        output_formula=template_cell.formula if template_cell else None,
                        dependencies=sorted(dependencies),
                        dependency_details=dependency_details,
                        source_locations=source_locations,
                        rule_reference="ASC 230",
                        review_status="Linked" if source_locations else "Review",
                    )
                )
        return links

    def _dependency_detail(self, dependency: str) -> dict[str, Any]:
        sheet_name, cell = dependency.split("!", 1)
        template_cell = self._template_cell(sheet_name, cell)
        if template_cell is not None:
            return {
                "key": dependency,
                "source_type": "generated",
                "value": self._get_value(sheet_name, cell),
                "formula": template_cell.formula,
            }
        return {
            "key": dependency,
            "source_type": "source_workbook",
            "actual_sheet": self.source.resolve_cell_sheet(sheet_name, cell),
            "actual_cell": self.source.actual_cell_name(sheet_name, cell),
            "value": self.source.value(sheet_name, cell),
            "formula": self.source.formula(sheet_name, cell),
        }

    def _source_locations_for_output(self, output_key: str) -> list[str]:
        locations: list[str] = []
        for dependency in sorted(self.evaluator.dependencies.get(output_key, set())):
            locations.extend(self._source_locations_for_dependency(dependency, set()))
        return _unique_locations(locations)

    def _source_locations_for_dependency(self, dependency: str, visited: set[str]) -> list[str]:
        if dependency in visited:
            return []
        visited.add(dependency)

        sheet_name, cell = dependency.split("!", 1)
        template_cell = self._template_cell(sheet_name, cell)
        if template_cell is None:
            detail = self._dependency_detail(dependency)
            if (
                detail.get("source_type") == "source_workbook"
                and detail.get("actual_sheet")
                and detail.get("actual_cell")
            ):
                return [self._format_source_location(detail)]
            return []

        source_locations: list[str] = []
        for nested_dependency in sorted(self.evaluator.dependencies.get(dependency, set())):
            source_locations.extend(self._source_locations_for_dependency(nested_dependency, visited))
        return source_locations

    def _write_workbook(
        self,
        path: Path,
        evidence: list[EvidenceLink],
        generated_values: dict[str, dict[str, Any]],
    ) -> None:
        wb = openpyxl.load_workbook(self.source.path)
        for sheet_name in GENERATED_SHEETS:
            template_sheet = self.template.sheets.get(sheet_name)
            if not template_sheet:
                continue
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                insert_at = 0 if sheet_name == "1. SCF" else 1
                ws = wb.create_sheet(sheet_name, insert_at)
            for cell, template_cell in template_sheet.cells.items():
                ws[cell] = template_cell.formula or template_cell.value

        evidence_rows = self._write_evidence_index(wb, evidence, generated_values)
        self._link_output_cells_to_evidence(wb, evidence, evidence_rows)
        wb.save(path)

    def _write_evidence_index(
        self,
        wb,
        evidence: list[EvidenceLink],
        generated_values: dict[str, dict[str, Any]],
    ) -> dict[str, int]:
        if EVIDENCE_INDEX_SHEET in wb.sheetnames:
            del wb[EVIDENCE_INDEX_SHEET]

        ws = wb.create_sheet(EVIDENCE_INDEX_SHEET, 0)
        headers = [
            "Output Sheet",
            "Output Cell",
            "Output Value",
            "Formula",
            "Dependencies",
            "Source Locations",
            "Rule Reference",
            "Review Status",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:H1"

        evidence_rows: dict[str, int] = {}
        for row_number, item in enumerate(evidence, start=2):
            output_key = f"{item.output_sheet}!{item.output_cell}"
            ws.append(
                [
                    item.output_sheet,
                    item.output_cell,
                    item.output_value,
                    item.output_formula or "",
                    "\n".join(item.dependencies),
                    "\n".join(item.source_locations),
                    item.rule_reference,
                    item.review_status,
                ]
            )
            evidence_rows[output_key] = row_number

        widths = {
            "A": 18,
            "B": 12,
            "C": 16,
            "D": 34,
            "E": 42,
            "F": 46,
            "G": 18,
            "H": 16,
        }
        for column, width in widths.items():
            ws.column_dimensions[column].width = width
        return evidence_rows

    def _link_output_cells_to_evidence(
        self,
        wb,
        evidence: list[EvidenceLink],
        evidence_rows: dict[str, int],
    ) -> None:
        for item in evidence:
            output_key = f"{item.output_sheet}!{item.output_cell}"
            row_number = evidence_rows.get(output_key)
            if not row_number or item.output_sheet not in wb.sheetnames:
                continue

            ws = wb[item.output_sheet]
            cell = ws[item.output_cell]
            cell.hyperlink = f"#'{EVIDENCE_INDEX_SHEET}'!A{row_number}"
            cell.font = Font(color="0563C1", underline="single")
            cell.comment = Comment(self._evidence_comment(item, row_number), "AI Reporting")

    def _evidence_comment(self, item: EvidenceLink, row_number: int) -> str:
        template_cell = self._template_cell(item.output_sheet, item.output_cell)
        source_text = "\n".join(item.source_locations[:5]) or "See dependency chain in Evidence Index."
        formula_text = template_cell.formula if template_cell and template_cell.formula else "Static/template value"
        return (
            f"Evidence Index row {row_number}\n"
            f"Output: {item.output_sheet}!{item.output_cell}\n"
            f"Formula: {formula_text}\n"
            f"Sources:\n{source_text}\n"
            f"Rule: {item.rule_reference}"
        )

    def _format_source_location(self, detail: dict[str, Any]) -> str:
        canonical = str(detail.get("key") or "")
        actual_sheet = detail.get("actual_sheet")
        actual_cell = detail.get("actual_cell")
        if actual_sheet and actual_cell:
            actual = f"{actual_sheet}!{actual_cell}"
            if actual != canonical:
                return f"{canonical} located at {actual}"
        return canonical


def _unique_locations(locations: list[str]) -> list[str]:
    seen = set()
    unique = []
    for location in locations:
        if location in seen:
            continue
        seen.add(location)
        unique.append(location)
    return unique
